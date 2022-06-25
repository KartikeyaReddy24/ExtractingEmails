from tld import get_tld
from openpyxl import load_workbook
from googlesearch import search
from tld import get_tld
import re
import requests
import pandas as pd
import time
import random
start_time = time.time()

LinksCount=0

#PhoneNumber=set()
SearchingFor=[]
LinksSearched=[]
l=set()
NotFoundLink=set()
links=set()
Nooflinkssearched=[]

should_restart = True

book = load_workbook('item1.xlsx')
sheet = book.active

while should_restart:
    should_restart = False
    for no in range(25):
        print(no)
        Nooflinkssearched.append(no)
        if LinksCount == 24:
            print("You are now at 2 minutes break")
            time.sleep(120)
            should_restart = True
            print("Total No of Web searches done: ",len(Nooflinkssearched))
            break

        for row in sheet.iter_rows():
            for cell in row:
                print("\nSearching for: ",cell.value)
                headers = {'User-Agent': 'Mozilla/5.0 (Windows NT 6.2; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/75.0.3770.100 Safari/537.36'}
                url = 'https://www.google.com?q=' + cell.value

                for i in search(cell.value, num_results=1):
                    print("\nFor loop started. Wait 30 seconds\n")
                    random.uniform(30,60)
                    time.sleep(30)
                    LinksCount +=1
                    print(f'Searching for: {i}\n\nNo. of Web Searches: {LinksCount}')
                    #print(i)
                    SearchingFor.append(cell.value)
                    LinksSearched.append(i)
                    try:
                        print("\nYou are at Extracting Emails. Wait for 18 seconds\n")
                        random.uniform(30,60)
                        time.sleep(18)
                        EMAIL_REGEX=r'[\w.+-]+@[\w-]+\.[\w.-]+'                 #r"[a-z0-9\.\-+_]+@[a-z0-9\.\-+_]+\.[a-z]+"         #r'(?:\.?)([\w\-_+#~!$&\'\.]+(?<!\.)(@|[ ]?\(?[ ]?(at|AT)[ ]?\)?[ ]?)(?<!\.)[\w]+[\w\-\.]*\.[a-zA-Z-]{2,3})(?:[^\w])'  #re.findall(r'[\w.+-]+@[\w-]+\.[\w.-]+')            # r"[\w\.-]+@[\w\.-]+"
                        #phone = r'\(?\b[2-9][0-9]{2}\)?[-][2-9][0-9]{2}[-][0-9]{4}\b'
                        r=requests.get(i, headers=headers)
                        for re_match in re.findall(EMAIL_REGEX, r.text):
                            l.add(re_match)
                            links.add(i)
                            #PhoneNumber.add(re_match)
                            print(re_match)
                    except:
                        print("Not Found: ",i)
                        NotFoundLink.add(i)
        exit
           

websearch=pd.DataFrame(SearchingFor)
searchlinks=pd.DataFrame(LinksSearched)
copying_links=pd.DataFrame(links)
data = pd.DataFrame(l)
NotFounddata = pd.DataFrame(NotFoundLink)
#PhNos=pd.DataFrame(PhoneNumber)
LinksEmails=pd.concat((copying_links, data), ignore_index=True, axis=1)
searchData=pd.concat((websearch, data), ignore_index=True, axis=1)


print("\nLinks Searched\n",copying_links)
print("\nEmail-ID's Collected:\n",data)
print("\nNot Found Links:\n",NotFounddata)
print("\nLinks Searched and Emails Collected:\n",LinksEmails)
#print("\nPhone Number Collected:\n",PhNos)

with pd.ExcelWriter('Industries.xlsx', engine='xlsxwriter') as writer:    
    # Write each dataframe to a different worksheet.
    searchData.to_excel(writer, sheet_name='Websearch Cell & Emails')
    searchlinks.to_excel(writer, sheet_name='LinksSearched')
    copying_links.to_excel(writer, sheet_name='UniqueLinksSearched')
    data.to_excel(writer, sheet_name='Emails')
    LinksEmails.to_excel(writer, sheet_name='Links & Emails')
    NotFounddata.to_excel(writer, sheet_name='NotFoundList')
    writer.save()

print("\n------------------------------- DATA SAVED -------------------------------\n")

print("\n----------- Time Take For Execution: %s seconds -----------\n" % (time.time() - start_time))