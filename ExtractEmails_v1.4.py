from tld import get_tld
from openpyxl import load_workbook
from googlesearch import search
from tld import get_tld
import re
import requests
import pandas as pd
import time
import random
import os

# Get environment variables
MINIMUM = os.getenv('min_val')
MAXIMUM = os.environ.get('max_val')

print("The Minimum row: ",MINIMUM)
print("The Maximum row: ",MAXIMUM)

start_time = time.time()

LinksCount=0

#PhoneNumber=set()
SearchingFor=[]
LinksSearched=[]
l=set()
NotFoundLink=set()
links=set()
Nooflinkssearched=[]



book = load_workbook('item.xlsx')
sheet = book.active
count = 0
# how to read this from env variables 

min_val=1
max_val=1

for row in sheet.iter_rows(min_row=min_val,max_row=max_val):
    for cell in row:
        print("\nSearching for: ",cell.value)
        headers = {'User-Agent': 'Mozilla/5.0 (Windows NT 6.2; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/75.0.3770.100 Safari/537.36'}
        search_value=''
        for val in cell.value.split():
            search_value+=val+'+'

        # url = 'https://www.google.com?q=' + search_value

        for i in search(search_value, num_results=1):
            #print("count ",count)
            if count>0 and count % 18 == 0:
                print(f'\n\n\t\t::::::::::: You have now reached maximun Websearches. Please wait :::::::::::\n\n')
                time.sleep(300)
            count+=1
            #print("\nFor loop started. Wait 30 seconds\n")
            random.uniform(30,60)
            time.sleep(30)
            LinksCount +=1
            print(f'\n\n\tNo. of Web Searches: {LinksCount}\n\tExtracting Emails in: {i}')
            #print(i)
            SearchingFor.append(cell.value)
            LinksSearched.append(i)
            try:
                #print("\nYou are at Extracting Emails. Wait for 18 seconds\n")
                random.uniform(30,60)
                time.sleep(18)
                EMAIL_REGEX= r"[a-z0-9\.\-+_]+@[a-z0-9\.\-+_]+\.[a-z]+"      #r'[\w.+-]+@[\w-]+\.[\w.-]+'        #r'(?:\.?)([\w\-_+#~!$&\'\.]+(?<!\.)(@|[ ]?\(?[ ]?(at|AT)[ ]?\)?[ ]?)(?<!\.)[\w]+[\w\-\.]*\.[a-zA-Z-]{2,3})(?:[^\w])'  #re.findall(r'[\w.+-]+@[\w-]+\.[\w.-]+')            # r"[\w\.-]+@[\w\.-]+"
                #phone = r'\(?\b[2-9][0-9]{2}\)?[-][2-9][0-9]{2}[-][0-9]{4}\b'
                r=requests.get(i, headers=headers)
                for re_match in re.findall(EMAIL_REGEX, r.text):
                    l.add(re_match)
                    links.add(i)
                    #PhoneNumber.add(re_match)
                    print(f'\n\t',re_match)
            except:
                print("\n\tNot Found: ",i)
                NotFoundLink.add(i)
           

websearch=pd.DataFrame(SearchingFor)
searchlinks=pd.DataFrame(LinksSearched)
copying_links=pd.DataFrame(links)
data = pd.DataFrame(l)
NotFounddata = pd.DataFrame(NotFoundLink)
#PhNos=pd.DataFrame(PhoneNumber)
LinksEmails=pd.concat((copying_links, data), ignore_index=True, axis=1)
searchData=pd.concat((websearch, data), ignore_index=True, axis=1)


print("\nLinks Searched\n",copying_links)
print("\nEmail-ID's Collected:\n",data)
print("\nNot Found Links:\n",NotFounddata)
print("\nLinks Searched and Emails Collected:\n",LinksEmails)
#print("\nPhone Number Collected:\n",PhNos)

with pd.ExcelWriter("Industries_"+str(MINIMUM)+"_"+str(MAXIMUM)+".xlsx", engine='xlsxwriter') as writer:    
    # Write each dataframe to a different worksheet.
    searchData.to_excel(writer, sheet_name='Websearch Cell & Emails')
    searchlinks.to_excel(writer, sheet_name='LinksSearched')
    copying_links.to_excel(writer, sheet_name='UniqueLinksSearched')
    data.to_excel(writer, sheet_name='Emails')
    LinksEmails.to_excel(writer, sheet_name='Links & Emails')
    NotFounddata.to_excel(writer, sheet_name='NotFoundList')
    writer.save()


def convert(seconds):
    seconds = seconds % (24 * 3600)
    hour = seconds // 3600
    seconds %= 3600
    minutes = seconds // 60
    seconds %= 60
      
    return "%d:%02d:%02d" % (hour, minutes, seconds)

print("\n------------------------------- DATA SAVED -------------------------------\n")

print("\n\t----------- Time Taken For Execution: %s  -----------\n\n" % (convert(time.time() - start_time)))
